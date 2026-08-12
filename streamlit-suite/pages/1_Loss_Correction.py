import io
import random
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st

from scipy.optimize import differential_evolution
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ==========================================================
# PAGE CONFIG
# ==========================================================

st.set_page_config(
    page_title="Loss Correction — Solar Suite",
    page_icon="⚡",
    layout="wide",
)


# ==========================================================
# CONSTANTS
# ==========================================================

QUOTES = [
    "☕ Vo kehte the kya ho tum, aaj hum kehte hai tum kya ho be?",
    "🌦 Aapka mann nahi kar raha bahar jaane ka?",
    "😊 Jinke ghar sheeshe ke bane hote hai vo basement mai kapde change krte h...",
    "😋 Aromatic Rose Latte with Frothy Milk pine ka mann hor hai na...",
    "🥛 Garmi mai daalo dudh mai Ice 🧊 Dudh bangya Very Nice...",
    "🌟 Aapke face pr toh Modiji se bhi jyda glow hai..",
    "😁 Horaha hai benstokes Kaan mai ghusjao insaan ke...",
    "😗 Muskuraiye aap MAL mai hai...",
    "🥱 Hum na hote toh Operations ka kya hota?",
    "😎 6:30 hote hi Billu MAL se faraar...",
    "😇 Guruji ne ek baat kahi thi....",
    "🎼 Karna hai kuchh kaam M se gaao...",
    "😠 Nahi karni Loss Correction, Now what to do?",
    "💸 Iss Job ko chhod or chhod kar ameer ho..",
]

MAX_DE_ITER = 60
DE_POPSIZE = 10


# ==========================================================
# CSS
# ==========================================================

st.markdown(
    """
    <style>

    div[data-testid="metric-container"] {
        background: #111827;
        border: 1px solid #1f2937;
        border-radius: 10px;
        padding: 12px 20px;
    }

    div[data-testid="stDataEditor"] {
        border-radius: 10px;
    }

    </style>
    """,
    unsafe_allow_html=True,
)


# ==========================================================
# SIDEBAR
# ==========================================================

st.sidebar.markdown(
    """
    <h1 style='text-align:center;
    background:linear-gradient(90deg,#00c6ff,#0072ff);
    -webkit-background-clip:text;
    -webkit-text-fill-color:transparent;
    font-size:40px;font-weight:800;'>
    ⚡ Solar Suite
    </h1>

    <p style='text-align:center;color:gray;font-size:14px;'>
    Forecast Correction Platform
    </p>
    """,
    unsafe_allow_html=True,
)

st.sidebar.divider()


# ==========================================================
# FILE READERS
# ==========================================================

@st.cache_data(show_spinner=False)
def read_excel_sheet(file_bytes, sheet_name, header=0, **kwargs):

    return pd.read_excel(
        io.BytesIO(file_bytes),
        sheet_name=sheet_name,
        header=header,
        **kwargs,
    )


@st.cache_data(show_spinner=False)
def get_sheet_names(file_bytes):

    xls = pd.ExcelFile(
        io.BytesIO(file_bytes)
    )

    return tuple(xls.sheet_names)


# ==========================================================
# WORKBOOK DETECTION
# ==========================================================

@st.cache_data(show_spinner=False)
def detect_workbook(file_bytes):

    sheets = get_sheet_names(
        file_bytes
    )

    is_cluster = "Fixed-CL1" in sheets

    calculation_sheet = (
        "Fixed-CL1"
        if is_cluster
        else "Fixed"
    )

    if is_cluster:

        ghi_cols = [
            "CL1-GHI",
            "CL2-GHI",
            "CL3-GHI",
            "CL4-GHI",
            "CL5-GHI",
        ]

    else:

        ghi_cols = [
            "GHI_Forecast"
        ]

    df = read_excel_sheet(
        file_bytes,
        calculation_sheet,
        header=1,
    )

    df.columns = [
        str(c).strip()
        for c in df.columns
    ]

    required = (
        ghi_cols
        + ["Actual", "Date"]
    )

    missing = [
        c
        for c in required
        if c not in df.columns
    ]

    if missing:

        raise ValueError(
            "Missing columns: "
            + ", ".join(missing)
        )

    df["Actual"] = pd.to_numeric(
        df["Actual"],
        errors="coerce",
    ).fillna(0)

    for col in ghi_cols:

        df[col] = pd.to_numeric(
            df[col],
            errors="coerce",
        ).fillna(0)

    # Remove rows after first blank date
    null_idx = df[
        df["Date"].isna()
    ].index

    if len(null_idx):

        first_null = null_idx[0]

        df = df.loc[
            : first_null - 1
        ]

    df = df.iloc[
        :96
    ].reset_index(
        drop=True
    )

    return (
        is_cluster,
        tuple(ghi_cols),
        df[
            ghi_cols + ["Actual"]
        ].copy(),
    )


# ==========================================================
# AREA & EFFICIENCY
# ==========================================================

@st.cache_data(show_spinner=False)
def read_area_efficiency(file_bytes):

    for header in [1, 2, 0]:

        try:

            df = read_excel_sheet(
                file_bytes,
                "Area & Efficiency",
                header=header,
                usecols=range(8),
            )

        except Exception:
            continue

        df.columns = [
            str(c).strip()
            for c in df.columns
        ]

        if (
            "Module Type"
            not in df.columns
        ):
            continue

        if (
            "Standard PV Efficiency (%)"
            not in df.columns
        ):
            continue

        if (
            "Total area(m2)"
            not in df.columns
        ):
            continue

        df = df[
            df["Module Type"].notna()
        ].copy()

        for col in df.columns:

            if col != "Module Type":

                df[col] = pd.to_numeric(
                    df[col],
                    errors="coerce",
                )

        df = df.dropna(
            subset=[
                "Standard PV Efficiency (%)",
                "Total area(m2)",
            ]
        )

        df = df[
            df[
                "Standard PV Efficiency (%)"
            ].between(1, 50)
        ]

        if len(df):

            return df.reset_index(
                drop=True
            )

    raise ValueError(
        "Could not correctly read "
        "Area & Efficiency sheet."
    )


# ==========================================================
# CLUSTER WEIGHTS
# ==========================================================

@st.cache_data(show_spinner=False)
def read_cluster_weights(file_bytes):

    df = read_excel_sheet(
        file_bytes,
        "Area & Efficiency",
        header=2,
        usecols=[12, 13, 14, 15, 16],
    )

    weights = []

    for i in range(1, 6):

        col = f"CL-{i}"

        if col not in df.columns:

            raise ValueError(
                f"{col} not found in "
                "Area & Efficiency."
            )

        weights.append(
            float(
                df[col].iloc[0]
            )
        )

    return tuple(weights)


# ==========================================================
# FORECAST CONFIG
# ==========================================================

@st.cache_data(show_spinner=False)
def read_latitude(file_bytes):

    df = read_excel_sheet(
        file_bytes,
        "Forecast Config",
        header=8,
    )

    if "Lat" not in df.columns:

        raise ValueError(
            "Lat column not found."
        )

    return float(
        df.loc[0, "Lat"]
    )


# ==========================================================
# TILT
# ==========================================================

@st.cache_data(show_spinner=False)
def read_tilt(file_bytes):

    df = read_excel_sheet(
        file_bytes,
        "Config Tilt Angle",
        header=7,
    )

    df.columns = [
        str(c).strip()
        for c in df.columns
    ]

    if "Fixed" not in df.columns:

        raise ValueError(
            "Fixed tilt column not found."
        )

    df["Fixed"] = pd.to_numeric(
        df["Fixed"],
        errors="coerce",
    ).fillna(0)

    df = df[
        df["Fixed"] != 0
    ].copy()

    df = df.dropna(
        how="all",
        axis=1,
    )

    df = df.rename(
        columns={
            "Unnamed: 2": "Month_Num",
            "Unnamed: 3": "Month",
        }
    )

    return df.set_index(
        "Month"
    )["Fixed"].to_dict()


# ==========================================================
# BACKEND CAL
# ==========================================================

@st.cache_data(show_spinner=False)
def read_backend_cal(
    file_bytes,
    sheet_name,
):

    return read_excel_sheet(
        file_bytes,
        sheet_name,
    )


# ==========================================================
# EFFICIENCY
# ==========================================================

def apply_efficiency(
    area_df,
    loss,
):

    df = area_df.copy()

    df["Efficiency Losses(%)"] = (
        float(loss)
    )

    df["Net Efficiency (%)"] = (
        df[
            "Standard PV Efficiency (%)"
        ]
        - float(loss)
    )

    df["Eff Area"] = (
        df["Total area(m2)"]
        * df["Net Efficiency (%)"]
        / 100
    )

    return df


def show_efficiency(
    df,
):

    cols = [
        "Module Type",
        "Standard PV Efficiency (%)",
        "Efficiency Losses(%)",
        "Net Efficiency (%)",
        "Total area(m2)",
    ]

    disp = df[
        [
            c
            for c in cols
            if c in df.columns
        ]
    ].copy()

    for col in disp.select_dtypes(
        include=np.number
    ).columns:

        disp[col] = disp[col].round(
            2
        )

    with st.expander(
        "🔍 View Efficiency Calculations"
    ):

        st.dataframe(
            disp,
            use_container_width=True,
            hide_index=True,
        )


# ==========================================================
# SOLAR GEOMETRY
# ==========================================================

@st.cache_data(show_spinner=False)
def solar_geometry(
    lat,
    tilt_dict=None,
    tracking=False,
):

    today = (
        pd.Timestamp.today()
        .normalize()
    )

    first = today.replace(
        month=1,
        day=1,
    )

    days = (
        today - first
    ).days + 1

    declination = (
        23.45
        * np.sin(
            np.radians(
                360
                * (284 + days)
                / 365
            )
        )
    )

    elevation = (
        90
        - lat
        + declination
    )

    if tracking:

        tilt = 0

    else:

        month = today.strftime(
            "%B"
        )

        tilt = (
            tilt_dict.get(
                month,
                0,
            )
            if tilt_dict
            else 0
        )

    sin_a = np.sin(
        np.radians(
            elevation
        )
    )

    sin_ab = np.sin(
        np.radians(
            elevation + tilt
        )
    )

    if abs(sin_a) < 1e-9:

        sin_a = 1e-9

    return sin_ab, sin_a


# ==========================================================
# LOSS OPTIMIZATION
# ==========================================================

@st.cache_data(show_spinner=False)
def optimize_loss(
    std_eff,
    area,
    actual,
    poa,
    cluster,
    weights,
):

    std_eff = np.asarray(
        std_eff,
        dtype=float,
    )

    area = np.asarray(
        area,
        dtype=float,
    )

    actual = np.asarray(
        actual,
        dtype=float,
    )

    peak = actual.max()

    if peak <= 0:

        return 0.0

    losses = np.arange(
        0,
        max(
            0.1,
            std_eff.min()
        ) + 0.01,
        0.1,
    )

    best_loss = 0
    best_error = np.inf

    if cluster:

        weights = np.asarray(
            weights,
            dtype=float,
        )

    for loss in losses:

        eff_area = (
            area
            * (
                std_eff - loss
            )
            / 100
        )

        if cluster:

            prediction = np.zeros(
                len(actual)
            )

            for i in range(
                min(5, len(eff_area))
            ):

                prediction += (
                    np.asarray(
                        poa[i],
                        dtype=float,
                    )
                    * eff_area[i]
                    * weights[i]
                    / 1e6
                )

        else:

            prediction = (
                np.asarray(
                    poa[0],
                    dtype=float,
                )
                * eff_area.sum()
                / 1e6
            )

        error = abs(
            peak
            - prediction.max()
        )

        if error < best_error:

            best_error = error
            best_loss = loss

    return float(
        round(best_loss, 2)
    )


# ==========================================================
# TRACKING FORECAST
# ==========================================================

def tracking_forecast(
    ghi_arrays,
    weights,
    blocks,
    dhi,
    start,
    end,
    maximum,
    east,
    west,
):

    blocks = np.asarray(
        blocks,
        dtype=float,
    )

    if (
        start >= maximum
        or maximum >= end
    ):

        return np.zeros(
            len(blocks)
        )

    m1 = 90 / (
        start - 1 - maximum
    )

    m2 = 90 / (
        end + 1 - maximum
    )

    zenith = np.where(
        blocks <= maximum,

        np.minimum(
            89,
            m1
            * (
                blocks
                - maximum
            ),
        ),

        np.minimum(
            89,
            m2
            * (
                blocks
                - maximum
            ),
        ),
    )

    panel = np.where(
        blocks < maximum,

        np.minimum(
            zenith,
            abs(east),
        ),

        np.where(
            (
                blocks > maximum
            )
            & (
                zenith > west
            ),

            west,
            zenith,
        ),
    )

    cos_a = np.clip(
        np.cos(
            np.radians(
                panel
            )
        ),
        1e-6,
        None,
    )

    forecast = np.zeros(
        len(blocks),
        dtype=float,
    )

    for ghi, weight in zip(
        ghi_arrays,
        weights,
    ):

        ghi = np.asarray(
            ghi,
            dtype=float,
        )

        dhi_value = (
            ghi
            * float(dhi)
            / 100
        )

        forecast += (
            (
                (
                    ghi
                    - dhi_value
                )
                / cos_a
            )
            * float(weight)
            / 1e6
        )

    return forecast


# ==========================================================
# DIFFERENTIAL EVOLUTION
# ==========================================================

def run_tracking_optimization(
    actual,
    blocks,
    ghi_arrays,
    weights,
):

    actual = np.asarray(
        actual,
        dtype=float,
    )

    blocks = np.asarray(
        blocks,
        dtype=float,
    )

    mask = (
        actual != 0
    )

    if not mask.any():

        raise ValueError(
            "Actual data contains "
            "no non-zero values."
        )

    actual_m = actual[
        mask
    ]

    bounds = [
        (0, 10),    # DHI
        (0, 30),    # Start
        (65, 80),   # End
        (44, 60),   # Maximum
        (0, 70),    # East
        (0, 70),    # West
    ]

    counter = {
        "generation": 0
    }

    progress = st.progress(
        0
    )

    status = st.empty()

    def objective(x):

        try:

            DHI = int(
                round(x[0])
            )

            start = int(
                round(x[1])
            )

            end = int(
                round(x[2])
            )

            maximum = int(
                round(x[3])
            )

            east = int(
                round(x[4])
            )

            west = int(
                round(x[5])
            )

            if (
                start >= maximum
                or maximum >= end
            ):

                return 1e9

            prediction = (
                tracking_forecast(
                    ghi_arrays,
                    weights,
                    blocks,
                    DHI,
                    start,
                    end,
                    maximum,
                    east,
                    west,
                )
            )

            prediction = (
                prediction[mask]
            )

            if (
                len(prediction) == 0
                or np.isnan(
                    prediction
                ).any()
                or np.isinf(
                    prediction
                ).any()
            ):

                return 1e9

            peak = actual_m.max()

            energy = actual_m.sum()

            if (
                peak <= 0
                or energy <= 0
            ):

                return 1e9

            block_error = (
                np.mean(
                    np.abs(
                        actual_m
                        - prediction
                    )
                )
                / peak
            )

            peak_error = (
                abs(
                    peak
                    - prediction.max()
                )
                / peak
            )

            energy_error = (
                abs(
                    energy
                    - prediction.sum()
                )
                / energy
            )

            return (
                0.80
                * block_error
                + 0.10
                * peak_error
                + 0.10
                * energy_error
            )

        except Exception:

            return 1e9

    def callback(
        xk,
        convergence,
    ):

        counter[
            "generation"
        ] += 1

        generation = (
            counter[
                "generation"
            ]
        )

        progress.progress(
            min(
                generation
                / MAX_DE_ITER,
                1.0,
            )
        )

        status.info(
            f"⚙️ Optimizing tracking parameters...\n\n"
            f"Generation "
            f"{generation} / {MAX_DE_ITER}"
        )

    try:

        result = (
            differential_evolution(
                objective,
                bounds=bounds,
                strategy="best1bin",
                maxiter=MAX_DE_ITER,
                popsize=DE_POPSIZE,
                tol=0.001,
                mutation=(0.5, 1),
                recombination=0.7,
                seed=42,
                polish=True,
                workers=1,
                callback=callback,
            )
        )

    finally:

        progress.empty()
        status.empty()

    return (
        np.round(
            result.x
        ).astype(int)
    )


# ==========================================================
# CHART
# ==========================================================

def make_chart(
    forecast,
    actual,
    title="Forecast vs Actual",
):

    forecast = np.asarray(
        forecast,
        dtype=float,
    )

    actual = np.asarray(
        actual,
        dtype=float,
    )

    x = np.arange(
        1,
        len(actual) + 1,
    )

    fig = go.Figure()

    fig.add_trace(
        go.Scatter(
            x=x,
            y=forecast,
            mode="lines",
            name="Forecast",
            line={
                "color": "#00c6ff",
                "width": 3,
            },
        )
    )

    fig.add_trace(
        go.Scatter(
            x=x,
            y=actual,
            mode="lines",
            name="Actual",
            line={
                "color": "#ef4444",
                "width": 3,
            },
        )
    )

    fig.update_layout(
        title=title,
        template="streamlit",
        height=500,
        hovermode="x unified",
        xaxis={
            "title": "15 Minute Block",
            "dtick": 4,
        },
        yaxis={
            "title": "Power (MW)",
        },
        legend={
            "orientation": "h",
            "y": 1.08,
            "x": 0,
        },
        margin={
            "l": 20,
            "r": 20,
            "t": 60,
            "b": 20,
        },
    )

    return fig


# ==========================================================
# FINAL REPORT
# ==========================================================

def style_sheet(ws):

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="0072FF",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in ws[1]:

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center"
        )

    for column in ws.columns:

        max_length = 0

        letter = (
            get_column_letter(
                column[0].column
            )
        )

        for cell in column:

            try:

                max_length = max(
                    max_length,
                    len(
                        str(
                            cell.value
                        )
                    ),
                )

            except Exception:
                pass

        ws.column_dimensions[
            letter
        ].width = min(
            max_length + 2,
            35,
        )


def build_report(
    original_bytes,
    result_df,
    parameters,
):

    output = io.BytesIO()

    output.write(
        original_bytes
    )

    output.seek(0)

    wb = load_workbook(
        output
    )

    if (
        "Final Report"
        in wb.sheetnames
    ):

        del wb[
            "Final Report"
        ]

    ws = wb.create_sheet(
        "Final Report"
    )

    # ------------------------------------------------------
    # TITLE
    # ------------------------------------------------------

    ws["A1"] = (
        "Aeromal Solar Suite - Final Report"
    )

    ws["A1"].font = Font(
        bold=True,
        size=16,
    )

    # ------------------------------------------------------
    # PARAMETERS
    # ------------------------------------------------------

    row = 3

    for key, value in (
        parameters.items()
    ):

        ws.cell(
            row=row,
            column=1,
            value=key,
        )

        ws.cell(
            row=row,
            column=2,
            value=value,
        )

        row += 1

    row += 1

    # ------------------------------------------------------
    # RESULT TABLE
    # ------------------------------------------------------

    headers = list(
        result_df.columns
    )

    for col_idx, column in enumerate(
        headers,
        start=1,
    ):

        ws.cell(
            row=row,
            column=col_idx,
            value=column,
        )

    header_row = row

    for values in result_df.itertuples(
        index=False,
        name=None,
    ):

        row += 1

        for col_idx, value in enumerate(
            values,
            start=1,
        ):

            if isinstance(
                value,
                (
                    np.integer,
                    np.floating,
                ),
            ):

                value = float(
                    value
                )

            ws.cell(
                row=row,
                column=col_idx,
                value=value,
            )

    for cell in ws[
        header_row
    ]:

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="0072FF",
        )

        cell.font = Font(
            color="FFFFFF",
            bold=True,
        )

    style_sheet(
        ws
    )

    wb.save(
        output
    )

    output.seek(0)

    return output.getvalue()


# ==========================================================
# PAGE
# ==========================================================

st.title(
    "Pakima Pakam Ravi, 3-4 Loss Correction kar chuke hai!! 😎"
)


# ==========================================================
# FILE UPLOAD
# ==========================================================

uploaded = st.file_uploader(
    "📂 Yaha Excel File Feko!!",
    type=["xlsx"],
    key="lc_uploader",
)


if uploaded is None:

    st.info(
        "Pehle File toh upload karo!!!"
    )

    st.stop()


file_bytes = (
    uploaded.getvalue()
)


# ==========================================================
# FILE SESSION RESET
# ==========================================================

file_id = (
    uploaded.name,
    len(file_bytes),
)

if (
    st.session_state.get(
        "lc_file_id"
    )
    != file_id
):

    keys_to_clear = [
        "lc_file_id",
        "lc_calculated",
        "lc_result",
        "lc_params",
        "lc_plant_type",
        "lc_edited_input",
    ]

    for key in keys_to_clear:

        st.session_state.pop(
            key,
            None,
        )

    st.session_state[
        "lc_file_id"
    ] = file_id


# ==========================================================
# DETECT WORKBOOK
# ==========================================================

try:

    (
        is_cluster,
        ghi_cols,
        input_df,
    ) = detect_workbook(
        file_bytes
    )

except Exception as e:

    st.error(
        f"Workbook read nahi ho paya: {e}"
    )

    st.stop()


# ==========================================================
# PLANT TYPE
# ==========================================================

st.subheader(
    "🏭 Plant Type"
)

plant_type = st.pills(
    "Select Plant Type",
    [
        "🏗️ Fixed",
        "🔄 Tracking",
    ],
    default=st.session_state.get(
        "lc_plant_type",
        "🏗️ Fixed",
    ),
)

# Save current selection
st.session_state[
    "lc_plant_type"
] = plant_type


# ==========================================================
# INPUT DATA
# ==========================================================

st.subheader(
    "📊 Input Data"
)

if is_cluster:

    st.caption(
        "Cluster plant detected. "
        "Enter/edit GHI for each cluster and Actual Power."
    )

else:

    st.caption(
        "Enter/edit GHI Forecast and Actual Power."
    )


# IMPORTANT:
# Only GHI + Actual are shown.
display_input = input_df[
    list(ghi_cols) + ["Actual"]
].copy()


for col in display_input.columns:

    display_input[col] = pd.to_numeric(
        display_input[col],
        errors="coerce",
    ).fillna(0)


edited_input = st.data_editor(
    display_input,
    use_container_width=True,
    hide_index=True,
    num_rows="fixed",
    key="lc_input_editor",
)


# ==========================================================
# MAIN CALCULATION BUTTON
# ==========================================================

st.divider()

calculate = st.button(
    "🚀 Calculate Loss Correction",
    use_container_width=True,
    type="primary",
)


# ==========================================================
# CALCULATION FUNCTION
# ==========================================================

def calculate_loss_correction(
    file_bytes,
    edited_input,
    plant_type,
    is_cluster,
    ghi_cols,
):

    # ------------------------------------------------------
    # Read backend data
    # ------------------------------------------------------

    area_df = read_area_efficiency(
        file_bytes
    )

    lat = read_latitude(
        file_bytes
    )

    if plant_type == "🏗️ Fixed":

        calculation_sheet = (
            "Fixed-CL1"
            if is_cluster
            else "Fixed"
        )

    else:

        calculation_sheet = (
            "Fixed-CL1"
            if is_cluster
            else "Fixed"
        )

    backend_df = read_excel_sheet(
        file_bytes,
        calculation_sheet,
        header=1,
    )

    backend_df.columns = [
        str(c).strip()
        for c in backend_df.columns
    ]

    backend_df = backend_df.iloc[
        :96
    ].copy()

    n = min(
        len(
            backend_df
        ),
        len(
            edited_input
        ),
    )

    backend_df = (
        backend_df
        .iloc[:n]
        .reset_index(drop=True)
    )

    # ------------------------------------------------------
    # Replace ONLY input values
    # ------------------------------------------------------

    for col in ghi_cols:

        backend_df[col] = pd.to_numeric(
            edited_input[col]
            .iloc[:n],
            errors="coerce",
        ).fillna(0).values

    backend_df["Actual"] = pd.to_numeric(
        edited_input["Actual"]
        .iloc[:n],
        errors="coerce",
    ).fillna(0).values

    actual = backend_df[
        "Actual"
    ].to_numpy(float)

    # ------------------------------------------------------
    # Fixed
    # ------------------------------------------------------

    if plant_type == "🏗️ Fixed":

        tilt_dict = read_tilt(
            file_bytes
        )

        sin_ab, sin_a = (
            solar_geometry(
                lat,
                tilt_dict,
                tracking=False,
            )
        )

        suffixes = (
            [
                "",
                "-CL2",
                "-CL3",
                "-CL4",
                "-CL5",
            ]
            if is_cluster
            else [""]
        )

        poa_list = []

        for col, suffix in zip(
            ghi_cols,
            suffixes,
        ):

            ghi = backend_df[
                col
            ].to_numpy(float)

            poa_list.append(
                ghi
                * sin_ab
                / sin_a
            )

        if is_cluster:

            weights = (
                read_cluster_weights(
                    file_bytes
                )
            )

        else:

            weights = None

        best_loss = optimize_loss(
            tuple(
                area_df[
                    "Standard PV Efficiency (%)"
                ]
            ),
            tuple(
                area_df[
                    "Total area(m2)"
                ]
            ),
            tuple(actual),
            tuple(
                tuple(x)
                for x in poa_list
            ),
            is_cluster,
            weights,
        )

        area_eff = apply_efficiency(
            area_df,
            best_loss,
        )

        if is_cluster:

            forecast = np.zeros(
                n
            )

            for i in range(5):

                forecast += (
                    poa_list[i]
                    * area_eff[
                        "Eff Area"
                    ].iloc[i]
                    * weights[i]
                    / 1e6
                )

        else:

            forecast = (
                poa_list[0]
                * area_eff[
                    "Eff Area"
                ].sum()
                / 1e6
            )

        return {
            "mode": "Fixed",
            "forecast": forecast,
            "actual": actual,
            "percentile": None,
            "profile": forecast,
            "sym_profile": None,
            "best_loss": best_loss,
            "area_eff": area_eff,
            "params": {
                "Plant Type": "Fixed",
                "Efficiency Loss (%)":
                    round(
                        best_loss,
                        2,
                    ),
            },
        }

    # ------------------------------------------------------
    # Tracking
    # ------------------------------------------------------

    sin_ab, sin_a = (
        solar_geometry(
            lat,
            tracking=True,
        )
    )

    ghi_arrays = [
        backend_df[
            col
        ].to_numpy(float)
        for col in ghi_cols
    ]

    if is_cluster:

        weights_raw = (
            read_cluster_weights(
                file_bytes
            )
        )

        poa_list = [
            ghi
            * sin_ab
            / sin_a
            for ghi in ghi_arrays
        ]

        best_loss = optimize_loss(
            tuple(
                area_df[
                    "Standard PV Efficiency (%)"
                ]
            ),
            tuple(
                area_df[
                    "Total area(m2)"
                ]
            ),
            tuple(actual),
            tuple(
                tuple(x)
                for x in poa_list
            ),
            True,
            weights_raw,
        )

        area_eff = apply_efficiency(
            area_df,
            best_loss,
        )

        backend_sheet = (
            "Backend Cal CL1"
        )

    else:

        weights_raw = None

        poa = (
            ghi_arrays[0]
            * sin_ab
            / sin_a
        )

        best_loss = optimize_loss(
            tuple(
                area_df[
                    "Standard PV Efficiency (%)"
                ]
            ),
            tuple(
                area_df[
                    "Total area(m2)"
                ]
            ),
            tuple(actual),
            (
                tuple(poa),
            ),
            False,
            None,
        )

        area_eff = apply_efficiency(
            area_df,
            best_loss,
        )

        backend_sheet = (
            "Backend Cal"
        )

    # ------------------------------------------------------
    # Backend blocks
    # ------------------------------------------------------

    block_df = read_backend_cal(
        file_bytes,
        backend_sheet,
    )

    blocks = block_df[
        "Block No."
    ].to_numpy(float)

    blocks = blocks[
        :n
    ]

    ghi_arrays = [
        x[:n]
        for x in ghi_arrays
    ]

    # ------------------------------------------------------
    # Tracking weights
    # ------------------------------------------------------

    if is_cluster:

        weights_eff = tuple(
            float(
                area_eff[
                    "Eff Area"
                ].iloc[i]
            )
            * weights_raw[i]
            for i in range(5)
        )

    else:

        weights_eff = (
            float(
                area_eff[
                    "Eff Area"
                ].sum()
            ),
        )

    # ------------------------------------------------------
    # Optimize tracking parameters
    # ------------------------------------------------------

    best = run_tracking_optimization(
        actual,
        blocks,
        ghi_arrays,
        weights_eff,
    )

    params = {
        "DHI": int(best[0]),
        "start": int(best[1]),
        "end": int(best[2]),
        "max": int(best[3]),
        "east": int(best[4]),
        "west": int(best[5]),
        "loss": float(best_loss),
    }

    # ------------------------------------------------------
    # Forecast
    # ------------------------------------------------------

    forecast = tracking_forecast(
        ghi_arrays,
        weights_eff,
        blocks,
        params["DHI"],
        params["start"],
        params["end"],
        params["max"],
        params["east"],
        params["west"],
    )

    return {
        "mode": "Tracking",
        "forecast": forecast,
        "actual": actual,
        "percentile": None,
        "profile": forecast,
        "sym_profile": None,
        "best_loss": best_loss,
        "area_eff": area_eff,
        "blocks": blocks,
        "params": {
            "Plant Type": "Tracking",
            "Efficiency Loss (%)":
                round(
                    best_loss,
                    2,
                ),
            "DHI (%)":
                params["DHI"],
            "Starting Block":
                params["start"],
            "Ending Block":
                params["end"],
            "Max Block":
                params["max"],
            "East Limit":
                params["east"],
            "West Limit":
                params["west"],
        },
    }


# ==========================================================
# RUN CALCULATION
# ==========================================================

if calculate:

    with st.spinner(
        "Calculation chal rahi hai... ⏳"
    ):

        try:

            result = (
                calculate_loss_correction(
                    file_bytes,
                    edited_input,
                    plant_type,
                    is_cluster,
                    ghi_cols,
                )
            )

            st.session_state[
                "lc_result"
            ] = result

            st.session_state[
                "lc_calculated"
            ] = True

        except Exception as e:

            st.session_state[
                "lc_calculated"
            ] = False

            st.error(
                f"Calculation failed: {e}"
            )

            st.stop()


# ==========================================================
# SHOW RESULTS
# ==========================================================

if not st.session_state.get(
    "lc_calculated",
    False,
):

    st.info(
        "Plant Type select karo, "
        "GHI/Actual verify karo aur "
        "**Calculate Loss Correction** dabao."
    )

    st.stop()


result = (
    st.session_state[
        "lc_result"
    ]
)


# ==========================================================
# RESULT HEADER
# ==========================================================

st.divider()

st.subheader(
    "📊 Loss Correction Result"
)

metric1, metric2 = st.columns(
    2
)

metric1.metric(
    "Plant Type",
    result["mode"],
)

metric2.metric(
    "Efficiency Loss",
    f'{result["best_loss"]:.2f}%',
)


# ==========================================================
# EFFICIENCY
# ==========================================================

show_efficiency(
    result["area_eff"]
)


# ==========================================================
# TRACKING PARAMETERS
# ==========================================================

if result["mode"] == "Tracking":

    st.subheader(
        "🔄 Tracking Parameters"
    )

    params = result[
        "params"
    ].copy()

    st.caption(
        "Parameters change karo aur "
        "Recalculate dabao."
    )

    with st.form(
        "tracking_parameter_form"
    ):

        c1, c2, c3 = st.columns(
            3
        )

        loss = c1.number_input(
            "Efficiency Loss (%)",
            min_value=0.0,
            max_value=50.0,
            step=0.1,
            value=float(
                params[
                    "Efficiency Loss (%)"
                ]
            ),
        )

        DHI = c2.number_input(
            "DHI (%)",
            min_value=0,
            max_value=100,
            step=1,
            value=int(
                params["DHI (%)"]
            ),
        )

        start = c3.number_input(
            "Starting Block",
            min_value=0,
            max_value=95,
            step=1,
            value=int(
                params[
                    "Starting Block"
                ]
            ),
        )

        c1, c2, c3 = st.columns(
            3
        )

        end = c1.number_input(
            "Ending Block",
            min_value=1,
            max_value=96,
            step=1,
            value=int(
                params[
                    "Ending Block"
                ]
            ),
        )

        maximum = c2.number_input(
            "Max Block",
            min_value=1,
            max_value=95,
            step=1,
            value=int(
                params[
                    "Max Block"
                ]
            ),
        )

        east = c3.number_input(
            "East Limit",
            min_value=0,
            max_value=90,
            step=1,
            value=int(
                params[
                    "East Limit"
                ]
            ),
        )

        west = st.number_input(
            "West Limit",
            min_value=0,
            max_value=90,
            step=1,
            value=int(
                params[
                    "West Limit"
                ]
            ),
        )

        recalculate = st.form_submit_button(
            "🔄 Recalculate Tracking Forecast",
            use_container_width=True,
            type="primary",
        )

    if recalculate:

        if (
            start >= maximum
            or maximum >= end
        ):

            st.error(
                "Starting Block < Max Block < Ending Block hona chahiye."
            )

        else:

            # Rebuild efficiency
            area_eff = apply_efficiency(
                read_area_efficiency(
                    file_bytes
                ),
                loss,
            )

            # Read workbook setup
            _, current_ghi_cols, _ = (
                detect_workbook(
                    file_bytes
                )
            )

            backend_sheet = (
                "Backend Cal CL1"
                if is_cluster
                else "Backend Cal"
            )

            block_df = read_backend_cal(
                file_bytes,
                backend_sheet,
            )

            blocks = (
                block_df[
                    "Block No."
                ]
                .to_numpy(float)
            )

            n = min(
                len(
                    edited_input
                ),
                len(blocks),
            )

            blocks = blocks[
                :n
            ]

            ghi_arrays = [
                pd.to_numeric(
                    edited_input[
                        col
                    ],
                    errors="coerce",
                )
                .fillna(0)
                .to_numpy(float)[
                    :n
                ]
                for col in current_ghi_cols
            ]

            if is_cluster:

                weights_raw = (
                    read_cluster_weights(
                        file_bytes
                    )
                )

                weights_eff = tuple(
                    float(
                        area_eff[
                            "Eff Area"
                        ].iloc[i]
                    )
                    * weights_raw[i]
                    for i in range(5)
                )

            else:

                weights_eff = (
                    float(
                        area_eff[
                            "Eff Area"
                        ].sum()
                    ),
                )

            forecast = tracking_forecast(
                ghi_arrays,
                weights_eff,
                blocks,
                DHI,
                start,
                end,
                maximum,
                east,
                west,
            )

            actual = (
                pd.to_numeric(
                    edited_input[
                        "Actual"
                    ],
                    errors="coerce",
                )
                .fillna(0)
                .to_numpy(float)[
                    :n
                ]
            )

            result["forecast"] = (
                forecast
            )

            result["actual"] = (
                actual
            )

            result["area_eff"] = (
                area_eff
            )

            result["best_loss"] = (
                loss
            )

            result["params"] = {
                "Plant Type": "Tracking",
                "Efficiency Loss (%)":
                    round(loss, 2),
                "DHI (%)": DHI,
                "Starting Block": start,
                "Ending Block": end,
                "Max Block": maximum,
                "East Limit": east,
                "West Limit": west,
            }

            st.session_state[
                "lc_result"
            ] = result

            st.success(
                "✅ Tracking forecast recalculated."
            )

            st.rerun()


# ==========================================================
# CHART
# ==========================================================

st.subheader(
    "📈 Forecast vs Actual"
)

st.plotly_chart(
    make_chart(
        result["forecast"],
        result["actual"],
        (
            "Tracking Forecast vs Actual"
            if result["mode"]
            == "Tracking"
            else
            "Fixed Forecast vs Actual"
        ),
    ),
    use_container_width=True,
)


# ==========================================================
# FINAL RESULT TABLE
# ==========================================================

st.subheader(
    "📋 Result"
)

result_table = pd.DataFrame(
    {
        "Block": np.arange(
            1,
            len(
                result[
                    "actual"
                ]
            )
            + 1,
        ),
        "Forecast": np.round(
            result[
                "forecast"
            ],
            4,
        ),
        "Actual": np.round(
            result[
                "actual"
            ],
            4,
        ),
    }
)

st.dataframe(
    result_table,
    use_container_width=True,
    hide_index=True,
)


# ==========================================================
# FINAL REPORT
# ==========================================================

st.divider()

st.subheader(
    "📥 Final Report"
)

# ----------------------------------------------------------
# Report columns
# ----------------------------------------------------------

report_df = result_table.copy()

# Percentile is included in the report.
# For this loss-correction workflow there isn't a separately
# calculated percentile curve, so use the input GHI percentile
# only when it is available. Otherwise leave it blank.

report_df.insert(
    1,
    "Percentile",
    np.nan,
)

# ----------------------------------------------------------
# Parameters
# ----------------------------------------------------------

report_parameters = (
    result["params"].copy()
)

report_parameters[
    "Rows"
] = len(
    report_df
)

report_parameters[
    "Cluster Plant"
] = (
    "Yes"
    if is_cluster
    else "No"
)

# ----------------------------------------------------------
# Build Excel
# ----------------------------------------------------------

try:

    report_bytes = build_report(
        original_bytes=file_bytes,
        result_df=report_df,
        parameters=report_parameters,
    )

    base_name = (
        uploaded.name
        .rsplit(".", 1)[0]
    )

    output_name = (
        f"{base_name}_Loss_Correction_Report.xlsx"
    )

    st.download_button(
        label="📥 Download Final Excel Report",
        data=report_bytes,
        file_name=output_name,
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        use_container_width=True,
        type="primary",
    )

except Exception as e:

    st.error(
        f"Report generate nahi ho payi: {e}"
    )
