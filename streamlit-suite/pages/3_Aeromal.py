import io
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st

from scipy.signal import savgol_filter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ==========================================================
# PAGE CONFIG
# ==========================================================

st.set_page_config(
    page_title="Aeromal — Solar Suite",
    page_icon="⚡",
    layout="wide",
)

AEROMAL_PASSWORD = "asdfghjkl;'"

VALID_COLUMNS = [
    "Green Gen-Meter (Developer)",
    "Green Gen-SCADA (Developer)",
    "SEMS",
    "Power (MW)",
]


# ==========================================================
# SIDEBAR
# ==========================================================

st.sidebar.markdown(
    """
    <h1 style='text-align:center;
    background:linear-gradient(90deg,#00c6ff,#0072ff);
    -webkit-background-clip:text;
    -webkit-text-fill-color:transparent;
    font-size:40px;font-weight:800;'>
    ⚡ Solar Suite
    </h1>

    <p style='text-align:center;color:gray;font-size:14px;'>
    Forecast Correction Platform
    </p>
    """,
    unsafe_allow_html=True,
)

st.sidebar.divider()


# ==========================================================
# AUTH
# ==========================================================

if "aeromal_auth" not in st.session_state:
    st.session_state.aeromal_auth = False

if st.session_state.aeromal_auth:

    if st.sidebar.button("🚪 Logout", use_container_width=True):
        st.session_state.aeromal_auth = False
        st.rerun()

else:

    st.title("🔒 Access bas bade logo ke paas hai")

    password = st.text_input(
        "Enter Password",
        type="password",
    )

    if st.button(
        "Login",
        type="primary",
        use_container_width=True,
    ):
        if password == AEROMAL_PASSWORD:
            st.session_state.aeromal_auth = True
            st.rerun()
        else:
            st.error("Incorrect Password")

    st.stop()


# ==========================================================
# FILE READERS
# ==========================================================

@st.cache_data(show_spinner=False)
def read_excel_file(file_bytes):
    """Read Excel workbook sheet names."""
    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    return xls.sheet_names


@st.cache_data(show_spinner=False)
def read_excel_sheet(file_bytes, sheet_name):
    """Read selected Excel sheet."""
    return pd.read_excel(
        io.BytesIO(file_bytes),
        sheet_name=sheet_name,
    )


@st.cache_data(show_spinner=False)
def read_csv_file(file_bytes):
    """Read CSV file."""
    return pd.read_csv(
        io.BytesIO(file_bytes)
    )


# ==========================================================
# 95TH PERCENTILE
# ==========================================================

@st.cache_data(show_spinner=False)
def cached_percentile(power_tuple, days):
    power = np.asarray(
        power_tuple,
        dtype=float,
    )

    profile = power.reshape(
        days,
        96,
    )

    return tuple(
        np.percentile(
            profile,
            95,
            axis=0,
        ).tolist()
    )


# ==========================================================
# BEST SYMMETRY SHIFT
# ==========================================================

@st.cache_data(show_spinner=False)
def cached_best_shift(profile_tuple):

    profile = np.asarray(
        profile_tuple,
        dtype=float,
    )

    best_error = np.inf
    best_shift = 0

    for shift in range(96):

        shifted = np.roll(
            profile,
            -shift,
        )

        symmetric = (
            profile + shifted[::-1]
        ) / 2

        error = np.sqrt(
            np.mean(
                (profile - symmetric) ** 2
            )
        )

        if error < best_error:
            best_error = error
            best_shift = shift

    return int(best_shift)


# ==========================================================
# NO CURTAILMENT PIPELINE
# ==========================================================

@st.cache_data(show_spinner=False)
def cached_no_curtailment(
    power_tuple,
    days,
    window,
    power_availability,
):

    percentile = np.asarray(
        cached_percentile(
            power_tuple,
            days,
        )
    )

    smooth = savgol_filter(
        percentile,
        window_length=window,
        polyorder=3,
    )

    best_shift = cached_best_shift(
        tuple(smooth.tolist())
    )

    shifted = np.roll(
        smooth,
        -best_shift,
    )

    symmetric = (
        0.50 * smooth
        + 0.50 * shifted[::-1]
    )

    # ------------------------------------------------------
    # Morning / evening blending
    # ------------------------------------------------------

    active = np.where(
        percentile > 0.1
    )[0]

    if len(active):

        start = active[0]
        end = active[-1]
        blend = 8

        end_start = min(
            start + 1 + blend,
            96,
        )

        if end_start > start + 1:

            idx = np.arange(
                start + 1,
                end_start,
            )

            w = np.linspace(
                1,
                0,
                len(idx),
            )

            symmetric[idx] = (
                w * percentile[idx]
                + (1 - w) * symmetric[idx]
            )

        begin = max(
            end - blend,
            0,
        )

        if end > begin:

            idx = np.arange(
                begin,
                end,
            )

            w = np.linspace(
                0,
                1,
                len(idx),
            )

            symmetric[idx] = (
                w * percentile[idx]
                + (1 - w) * symmetric[idx]
            )

    # ------------------------------------------------------
    # Final smoothing
    # ------------------------------------------------------

    smooth = savgol_filter(
        percentile,
        window_length=11,
        polyorder=3,
    )

    symmetric = savgol_filter(
        symmetric,
        window_length=11,
        polyorder=3,
    )

    smooth = np.clip(
        smooth,
        0,
        None,
    )

    symmetric = np.clip(
        symmetric,
        0,
        None,
    )

    smooth[
        smooth < 0.1
    ] = 0

    symmetric[
        symmetric < 0.1
    ] = 0

    factor = (
        power_availability / 100
    )

    smooth *= factor
    symmetric *= factor

    return (
        tuple(percentile.tolist()),
        tuple(smooth.tolist()),
        tuple(symmetric.tolist()),
        best_shift,
    )


# ==========================================================
# MANUAL SHIFT
# ==========================================================

@st.cache_data(show_spinner=False)
def cached_sym_shift_nc(
    smooth_tuple,
    shift,
):

    smooth = np.asarray(
        smooth_tuple,
        dtype=float,
    )

    shifted = np.roll(
        smooth,
        -int(shift),
    )

    symmetric = (
        0.50 * smooth
        + 0.50 * shifted[::-1]
    )

    symmetric = savgol_filter(
        symmetric,
        window_length=11,
        polyorder=3,
    )

    symmetric = np.clip(
        symmetric,
        0,
        None,
    )

    symmetric[
        symmetric < 0.1
    ] = 0

    return tuple(
        symmetric.tolist()
    )


# ==========================================================
# CURTAILMENT PIPELINE
# ==========================================================

@st.cache_data(show_spinner=False)
def cached_curtailment(
    power_tuple,
    days,
    peak_cap,
    target_width,
    window,
    power_availability,
):

    percentile = (
        np.asarray(
            cached_percentile(
                power_tuple,
                days,
            )
        )
        * 1.03
    )

    y = percentile.copy()
    n = len(y)

    result = np.zeros(n)

    # ------------------------------------------------------
    # Smooth and gradient
    # ------------------------------------------------------

    smooth = savgol_filter(
        y,
        7,
        2,
    )

    gradient = np.gradient(
        smooth
    )

    # ------------------------------------------------------
    # Left side
    # ------------------------------------------------------

    left_peak = np.argmax(
        smooth[: n // 2]
    )

    if left_peak < 2:
        left_peak = 2

    left_start = np.argmax(
        gradient[:left_peak]
    )

    if left_start >= left_peak:
        left_start = 0

    x_left = np.arange(
        left_start,
        left_peak,
    )

    y_left = smooth[
        left_start:left_peak
    ]

    if len(x_left) < 2:
        m1, c1 = 1, 0
    else:
        m1, c1 = np.polyfit(
            x_left,
            y_left,
            1,
        )

    # ------------------------------------------------------
    # Right side
    # ------------------------------------------------------

    right_peak = (
        np.argmax(
            smooth[n // 2:]
        )
        + n // 2
    )

    threshold = (
        0.02 * np.max(smooth)
    )

    active = np.where(
        smooth > threshold
    )[0]

    if len(active):
        right_end = active[-1]
    else:
        right_end = n - 1

    if right_end <= right_peak:
        right_end = min(
            n - 1,
            right_peak + 2,
        )

    x_right = np.arange(
        right_peak,
        right_end,
    )

    y_right = smooth[
        right_peak:right_end
    ]

    if len(x_right) < 2:
        m2, c2 = -1, smooth[right_peak]
    else:
        m2, c2 = np.polyfit(
            x_right,
            y_right,
            1,
        )

    # ------------------------------------------------------
    # Trip point
    # ------------------------------------------------------

    if abs(m1) < 1e-9:
        m1 = 1e-9

    if abs(m2) < 1e-9:
        m2 = -1e-9

    A = (
        1 / m2
        - 1 / m1
    )

    B = (
        c1 / m1
        - c2 / m2
    )

    if abs(A) < 1e-9:
        trip = float(
            target_width
        )
    else:
        trip = max(
            0,
            (target_width - B) / A,
        )

    left_idx = int(
        np.clip(
            round((trip - c1) / m1),
            0,
            n - 1,
        )
    )

    right_idx = int(
        np.clip(
            round((trip - c2) / m2),
            0,
            n - 1,
        )
    )

    if right_idx <= left_idx:
        right_idx = min(
            n - 1,
            left_idx + 1,
        )

    # ------------------------------------------------------
    # Dome mode
    # ------------------------------------------------------

    if peak_cap >= trip:

        left_curve = np.zeros(n)

        for i in range(n):

            value = (
                m1 * i + c1
            )

            left_curve[i] = min(
                value,
                trip,
            )

            if value >= trip:
                left_idx = i
                break

        right_curve = np.zeros(n)

        for i in range(
            n - 1,
            -1,
            -1,
        ):

            value = (
                m2 * i + c2
            )

            right_curve[i] = min(
                value,
                trip,
            )

            if value >= trip:
                right_idx = i
                break

        result = np.maximum(
            left_curve,
            right_curve,
        )

        width = max(
            right_idx - left_idx,
            2,
        )

        dome_height = max(
            20,
            0.12 * trip,
        )

        xc = np.linspace(
            -1,
            1,
            width,
        )

        shape = np.sqrt(
            np.maximum(
                0,
                1 - xc ** 2,
            )
        )

        dome = (
            trip
            + dome_height * shape
        )

        dome[0] = trip
        dome[-1] = trip

        end_idx = min(
            left_idx + width,
            n,
        )

        result[
            left_idx:end_idx
        ] = dome[
            :end_idx-left_idx
        ]

        result = savgol_filter(
            result,
            window,
            3,
        )

        result = np.clip(
            result,
            0,
            None,
        )

        result[:left_start] = smooth[
            :left_start
        ]

        result[right_end:] = smooth[
            right_end:
        ]

        result = savgol_filter(
            result,
            7,
            3,
        )

        result *= (
            power_availability / 100
        )

        result = np.clip(
            result,
            0,
            None,
        )

        result[
            result < 0.2
        ] = 0

    # ------------------------------------------------------
    # Flat peak mode
    # ------------------------------------------------------

    else:

        for i in range(n):

            value = (
                m1 * i + c1
            )

            result[i] = value

            if value >= peak_cap:
                left_idx = i
                break

        right_curve = np.zeros(n)

        for i in range(
            n - 1,
            -1,
            -1,
        ):

            value = (
                m2 * i + c2
            )

            right_curve[i] = value

            if value >= peak_cap:
                right_idx = i
                break

        result = np.maximum(
            result,
            right_curve,
        )

        if right_idx > left_idx:

            result[
                left_idx:right_idx
            ] = peak_cap

        result = np.clip(
            result,
            0,
            peak_cap,
        )

        result = savgol_filter(
            result,
            window,
            3,
        )

        result[:left_start] = smooth[
            :left_start
        ]

        result[right_end:] = smooth[
            right_end:
        ]

        result = savgol_filter(
            result,
            7,
            3,
        )

        result *= (
            power_availability / 100
        )

        result = np.clip(
            result,
            0,
            peak_cap,
        )

        result[
            result < 1
        ] = 0

    best_shift = cached_best_shift(
        tuple(result.tolist())
    )

    return (
        tuple(percentile.tolist()),
        tuple(result.tolist()),
        best_shift,
    )


# ==========================================================
# CURTAILMENT SYMMETRY
# ==========================================================

@st.cache_data(show_spinner=False)
def cached_sym_shift_c(
    final_smooth_tuple,
    shift,
):

    smooth = np.asarray(
        final_smooth_tuple,
        dtype=float,
    )

    shifted = np.roll(
        smooth,
        -int(shift),
    )

    symmetric = (
        smooth + shifted[::-1]
    ) / 2

    return tuple(
        symmetric.tolist()
    )


# ==========================================================
# REPORT CREATION
# ==========================================================

def style_report_sheet(ws):

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="0072FF",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in ws[1]:

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center"
        )

    for column in ws.columns:

        max_length = 0
        column_letter = get_column_letter(
            column[0].column
        )

        for cell in column:

            try:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )
            except Exception:
                pass

        ws.column_dimensions[
            column_letter
        ].width = min(
            max_length + 2,
            35,
        )


def build_excel_report(
    original_bytes,
    file_type,
    source_df,
    selected_column,
    mode,
    result_df,
    parameters,
):

    output = io.BytesIO()

    # ------------------------------------------------------
    # Excel input
    # ------------------------------------------------------

    if file_type == "xlsx":

        output.write(
            original_bytes
        )
        output.seek(0)

        wb = load_workbook(
            output
        )

        # Remove previous Final Report
        if "Final Report" in wb.sheetnames:
            del wb["Final Report"]

        ws = wb.create_sheet(
            "Final Report"
        )

        # Report title
        ws["A1"] = (
            "Aeromal Solar Suite - Final Report"
        )

        ws["A1"].font = Font(
            bold=True,
            size=16,
        )

        ws["A3"] = "Selected Column"
        ws["B3"] = selected_column

        ws["A4"] = "Mode"
        ws["B4"] = mode

        row = 6

        # Parameters
        ws.cell(
            row=row,
            column=1,
            value="Parameters",
        ).font = Font(
            bold=True,
            size=12,
        )

        row += 1

        for key, value in parameters.items():

            ws.cell(
                row=row,
                column=1,
                value=key,
            )

            ws.cell(
                row=row,
                column=2,
                value=value,
            )

            row += 1

        row += 1

        # Result table
        for col_idx, column in enumerate(
            result_df.columns,
            start=1,
        ):

            ws.cell(
                row=row,
                column=col_idx,
                value=column,
            )

        header_row = row

        for r_idx, values in enumerate(
            result_df.itertuples(
                index=False,
                name=None,
            ),
            start=row + 1,
        ):

            for c_idx, value in enumerate(
                values,
                start=1,
            ):

                ws.cell(
                    row=r_idx,
                    column=c_idx,
                    value=float(value)
                    if isinstance(
                        value,
                        (np.floating, np.integer),
                    )
                    else value,
                )

        # Style result header
        for cell in ws[header_row]:

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="0072FF",
            )

            cell.font = Font(
                color="FFFFFF",
                bold=True,
            )

        style_report_sheet(
            ws
        )

        wb.save(
            output
        )

        output.seek(0)

        return output.getvalue()

    # ------------------------------------------------------
    # CSV input
    # ------------------------------------------------------

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:

        source_df.to_excel(
            writer,
            sheet_name="Source Data",
            index=False,
        )

        result_df.to_excel(
            writer,
            sheet_name="Final Report",
            index=False,
        )

    output.seek(0)

    wb = load_workbook(
        output
    )

    ws = wb["Final Report"]

    # Add metadata above report
    ws.insert_rows(1, 5)

    ws["A1"] = (
        "Aeromal Solar Suite - Final Report"
    )

    ws["A1"].font = Font(
        bold=True,
        size=16,
    )

    ws["A3"] = "Selected Column"
    ws["B3"] = selected_column

    ws["A4"] = "Mode"
    ws["B4"] = mode

    row = 6

    for key, value in parameters.items():

        ws.cell(
            row=row,
            column=1,
            value=key,
        )

        ws.cell(
            row=row,
            column=2,
            value=value,
        )

        row += 1

    style_report_sheet(
        ws
    )

    wb.save(
        output
    )

    output.seek(0)

    return output.getvalue()


# ==========================================================
# CHART
# ==========================================================

def make_chart(
    x,
    curves,
    mode,
):

    fig = go.Figure()

    for name, values, color in curves:

        fig.add_trace(
            go.Scatter(
                x=x,
                y=values,
                name=name,
                line=dict(
                    color=color,
                    width=4,
                ),
            )
        )

    fig.update_layout(
        title=f"Aeromal - {mode}",
        height=550,
        hovermode="x unified",
        template="streamlit",
        xaxis_title="Block",
        yaxis_title="Power",
        legend=dict(
            orientation="h",
            y=1.08,
            x=0,
        ),
        margin=dict(
            l=20,
            r=20,
            t=60,
            b=20,
        ),
    )

    return fig


# ==========================================================
# PAGE
# ==========================================================

st.title(
    "Kaha hai Aeromal ka khauf?!!"
)


# ==========================================================
# FILE UPLOADER
# ==========================================================

uploaded_file = st.file_uploader(
    "📂 Upload CSV or Excel file",
    type=["csv", "xlsx", "xls"],
)


if uploaded_file is None:

    st.info(
        "Pehle CSV ya Excel file upload karo!!!"
    )

    st.stop()


file_bytes = uploaded_file.getvalue()

file_name = uploaded_file.name.lower()


# ==========================================================
# LOAD DATA
# ==========================================================

if file_name.endswith(".csv"):

    file_type = "csv"

    try:

        source_df = read_csv_file(
            file_bytes
        )

    except Exception as e:

        st.error(
            f"CSV read nahi ho payi: {e}"
        )

        st.stop()

    sheet_name = None

else:

    file_type = "xlsx"

    try:

        sheet_names = read_excel_file(
            file_bytes
        )

    except Exception as e:

        st.error(
            f"Excel read nahi ho payi: {e}"
        )

        st.stop()

    sheet_name = st.selectbox(
        "📑 Select Excel Sheet",
        sheet_names,
    )

    source_df = read_excel_sheet(
        file_bytes,
        sheet_name,
    )


# ==========================================================
# CLEAN COLUMN NAMES
# ==========================================================

source_df.columns = [
    str(col).strip()
    for col in source_df.columns
]


# ==========================================================
# COLUMN SELECTION
# ==========================================================

available_columns = [
    col
    for col in VALID_COLUMNS
    if col in source_df.columns
]


if not available_columns:

    st.error(
        "Required columns nahi mili. "
        "File mein inmein se koi column hona chahiye:"
    )

    st.write(
        VALID_COLUMNS
    )

    st.write(
        "Available columns:"
    )

    st.write(
        list(source_df.columns)
    )

    st.stop()


st.subheader(
    "⚙️ Calculation Input"
)

selected_column = st.selectbox(
    "Select Generation / Power Column",
    available_columns,
)


# ==========================================================
# PREPARE POWER DATA
# ==========================================================

power_series = pd.to_numeric(
    source_df[selected_column],
    errors="coerce",
).fillna(0)


power_series = power_series[
    np.isfinite(
        power_series
    )
]


power_list = power_series.tolist()


if len(power_list) == 0:

    st.warning(
        "Selected column mein valid numerical data nahi hai."
    )

    st.stop()


# ==========================================================
# 96 BLOCK VALIDATION
# ==========================================================

n_rows = len(power_list)


if n_rows % 96 != 0:

    st.error(
        f"""
        Number of rows must be divisible by 96.

        Current rows: {n_rows}

        Required:
        96 rows = 1 day
        192 rows = 2 days
        288 rows = 3 days
        """
    )

    st.stop()


days = n_rows // 96


st.success(
    f"✅ {n_rows} rows detected = {days} day(s)"
)


power_tuple = tuple(
    float(x)
    for x in power_list
)


# ==========================================================
# MODE
# ==========================================================

st.markdown(
    """
    <style>
    div[data-testid="stToggle"]{
        background:#1f2937;
        border:2px solid #0072ff;
        border-radius:14px;
        padding:14px 18px;
        margin-bottom:15px;
    }

    div[data-testid="stToggle"] label{
        font-size:18px !important;
        font-weight:700 !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


curtailment = st.toggle(
    "⚡ Curtailment Mode",
    value=False,
)


# ==========================================================
# COMMON X AXIS
# ==========================================================

x = np.arange(
    1,
    97,
)


# ==========================================================
# NO CURTAILMENT
# ==========================================================

if not curtailment:

    st.subheader(
        "☀️ No Curtailment Mode"
    )

    col1, col2, col3 = st.columns(
        3
    )

    window = col1.number_input(
        "Window Length",
        min_value=5,
        max_value=31,
        step=2,
        value=11,
    )

    power_availability = col2.number_input(
        "Power Availability (%)",
        min_value=0,
        max_value=1000,
        value=100,
    )

    # Get calculated profile
    (
        ap_t,
        smooth_t,
        sym_t,
        best_shift,
    ) = cached_no_curtailment(
        power_tuple,
        days,
        window,
        power_availability,
    )

    # Manual shift
    shift = col3.number_input(
        "Shift",
        min_value=0,
        max_value=95,
        value=best_shift,
        step=1,
    )

    if shift != best_shift:

        sym_t = cached_sym_shift_nc(
            smooth_t,
            shift,
        )

    percentile = np.asarray(
        ap_t
    )

    smooth = np.asarray(
        smooth_t
    )

    symmetric = np.asarray(
        sym_t
    )

    # ------------------------------------------------------
    # Chart
    # ------------------------------------------------------

    fig = make_chart(
        x,
        [
            (
                "Sym Profile",
                symmetric,
                "#00c6ff",
            ),
            (
                "Profile",
                smooth,
                "#22c55e",
            ),
            (
                "95th Percentile",
                percentile,
                "#ef4444",
            ),
        ],
        "No Curtailment",
    )

    st.plotly_chart(
        fig,
        use_container_width=True,
    )

    # ------------------------------------------------------
    # Result table
    # ------------------------------------------------------

    result_df = pd.DataFrame(
        {
            "Block": x,
            "Percentile": percentile,
            "Profile": smooth,
            "Sym Profile": symmetric,
        }
    )

    st.subheader(
        "📊 Generated Curve"
    )

    st.dataframe(
        result_df,
        use_container_width=True,
        hide_index=True,
    )

    parameters = {
        "Selected Column": selected_column,
        "Mode": "No Curtailment",
        "Days": days,
        "Window Length": window,
        "Power Availability (%)": power_availability,
        "Best Shift": best_shift,
        "Applied Shift": shift,
    }


# ==========================================================
# CURTAILMENT
# ==========================================================

else:

    st.subheader(
        "⚡ Curtailment Mode"
    )

    if not np.any(
        np.asarray(power_list) > 0
    ):

        st.warning(
            "Please enter Power values to continue."
        )

        st.stop()

    col1, col2, col3 = st.columns(
        3
    )

    power_availability = col1.number_input(
        "Power Availability (%)",
        min_value=0,
        max_value=1000,
        value=100,
        step=1,
    )

    peak_cap = col1.number_input(
        "Peak Cap",
        min_value=1,
        value=max(
            1,
            int(max(power_list)),
        ),
        step=1,
    )

    target_width = col2.number_input(
        "Target Width",
        min_value=1,
        value=25,
        step=1,
    )

    window = col2.slider(
        "Window Length",
        min_value=5,
        max_value=31,
        value=11,
        step=2,
    )

    (
        ap_t,
        final_t,
        best_shift,
    ) = cached_curtailment(
        power_tuple,
        days,
        peak_cap,
        target_width,
        window,
        power_availability,
    )

    shift = col3.number_input(
        "Shift",
        min_value=0,
        max_value=95,
        value=best_shift,
        step=1,
    )

    symmetric_t = cached_sym_shift_c(
        final_t,
        shift,
    )

    percentile = np.asarray(
        ap_t
    )

    final_profile = np.asarray(
        final_t
    )

    symmetric = np.asarray(
        symmetric_t
    )

    # ------------------------------------------------------
    # Chart
    # ------------------------------------------------------

    fig = make_chart(
        x,
        [
            (
                "Generation",
                percentile,
                "#ef4444",
            ),
            (
                "Profile",
                final_profile,
                "#00c6ff",
            ),
            (
                "Sym Profile",
                symmetric,
                "#0072ff",
            ),
        ],
        "Curtailment",
    )

    st.plotly_chart(
        fig,
        use_container_width=True,
    )

    # ------------------------------------------------------
    # Result table
    # ------------------------------------------------------

    result_df = pd.DataFrame(
        {
            "Block": x,
            "Power": percentile,
            "Profile": final_profile,
            "Sym Profile": symmetric,
        }
    )

    st.subheader(
        "📊 Generated Curve"
    )

    st.dataframe(
        result_df,
        use_container_width=True,
        hide_index=True,
    )

    parameters = {
        "Selected Column": selected_column,
        "Mode": "Curtailment",
        "Days": days,
        "Power Availability (%)": power_availability,
        "Peak Cap": peak_cap,
        "Target Width": target_width,
        "Window Length": window,
        "Best Shift": best_shift,
        "Applied Shift": shift,
    }


# ==========================================================
# FINAL REPORT
# ==========================================================

st.divider()

st.subheader(
    "📥 Final Report"
)

st.caption(
    "The report will be added to the uploaded workbook "
    "as a new 'Final Report' sheet."
)


# ==========================================================
# SOURCE DATA FOR CSV REPORT
# ==========================================================

if file_type == "csv":

    report_source = source_df.copy()

else:

    report_source = source_df.copy()


# ==========================================================
# BUILD DOWNLOAD FILE
# ==========================================================

try:

    report_bytes = build_excel_report(
        original_bytes=file_bytes,
        file_type=file_type,
        source_df=report_source,
        selected_column=selected_column,
        mode=(
            "Curtailment"
            if curtailment
            else "No Curtailment"
        ),
        result_df=result_df,
        parameters=parameters,
    )

    base_name = (
        uploaded_file.name
        .rsplit(".", 1)[0]
    )

    output_name = (
        f"{base_name}_Aeromal_Report.xlsx"
    )

    st.download_button(
        label="📥 Download Final Excel Report",
        data=report_bytes,
        file_name=output_name,
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        use_container_width=True,
        type="primary",
    )

except Exception as e:

    st.error(
        f"Report generate nahi ho payi: {e}"
    )
